import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

def create_sales_order_excel(order_data: dict) -> bytes:
    wb = openpyxl.load_workbook("backend/templates/sales_order_template.xlsx")
    ws = wb["出貨單˙收據"]
    
    # Date parsing
    date_val = order_data.get('date', '')
    if isinstance(date_val, str) and date_val:
        try:
            date_str = datetime.fromisoformat(date_val.replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            date_str = date_val
    elif isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = ''
        
    cust_name = order_data.get('customer_name') or ''
    plate_num = order_data.get('plate_number') or ''
    created_by = order_data.get('created_by') or ''
    cust_addr = order_data.get('customer_address') or ''
    
    # Left side general info
    ws['D31'] = cust_name
    ws['D32'] = plate_num
    ws['J31'] = date_str
    ws['G32'] = created_by
    ws['D34'] = cust_addr
    
    # Right side general info
    ws['R31'] = cust_name
    ws['R32'] = plate_num
    ws['X31'] = date_str
    ws['U32'] = created_by
    ws['R34'] = cust_addr
    
    items = order_data.get('items', [])
    for idx, item in enumerate(items):
        if idx >= 8:  # Template supports up to 8 rows natively
            break
        r = 36 + idx
        brand = item.get('brand') or ''
        spec = item.get('spec') or item.get('specification') or ''
        pattern = item.get('pattern') or ''
        brand_spec_pattern = f"{brand} {spec} {pattern}".strip()
        service_type = item.get('service_type') or ''
        tire_pos = item.get('tire_position') or ''
        qty = item.get('qty') or 0
        price = item.get('price') or 0
        amount = qty * price
        note = item.get('note') or ''
        
        # Left side item row
        ws[f'B{r}'] = idx + 1
        ws[f'C{r}'] = order_data.get('order_id') or ''
        ws[f'D{r}'] = service_type
        ws[f'E{r}'] = brand_spec_pattern
        ws[f'G{r}'] = tire_pos
        ws[f'H{r}'] = note
        ws[f'I{r}'] = qty
        ws[f'J{r}'] = price
        ws[f'K{r}'] = amount
        
        # Right side item row
        ws[f'P{r}'] = idx + 1
        ws[f'Q{r}'] = order_data.get('order_id') or ''
        ws[f'R{r}'] = service_type
        ws[f'S{r}'] = brand_spec_pattern
        ws[f'U{r}'] = tire_pos
        ws[f'V{r}'] = ''
        ws[f'W{r}'] = qty
        ws[f'X{r}'] = price
        ws[f'Y{r}'] = amount

    # Totals
    subtotal = order_data.get('total_amount') or 0
    tax = order_data.get('tax_amount') or 0
    grand = order_data.get('grand_total') or 0
    
    # Left totals
    ws['K19'] = subtotal
    ws['H20'] = tax
    ws['K20'] = grand
    
    # Right totals
    ws['Y19'] = subtotal
    ws['V20'] = tax
    ws['Y20'] = grand
    
    # Save to memory
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_envelope_excel(party_data: dict, is_return: bool = False) -> bytes:
    wb = openpyxl.load_workbook("backend/templates/envelope_template.xlsx")
    ws = wb["信封袋列印"]
    
    is_vendor = "vendor_id" in party_data
    
    # Extract zip code
    addr = party_data.get('address') or ''
    zip_code = ''
    if addr:
        import re
        m = re.search(r'\d{3}', addr)
        if m:
            zip_code = m.group(0)
        else:
            zip_code = addr[:3] if addr[:3].isdigit() else ''

    if not is_vendor:
        ws['K6'] = ''
        ws['L6'] = ''
        ws['L8'] = ''
        ws['M8'] = ''
        
        if not is_return:
            ws['C6'] = zip_code
            ws['D6'] = addr
            ws['D8'] = party_data.get('name') or ''
            ws['E8'] = "會計部 收"
        else:
            ws['B5'] = f"\n         {party_data.get('name') or ''}\n\n      {addr}"
            ws['C6'] = "350"
            ws['D6'] = "苗栗縣竹南鎮中美里10鄰保安林52之10號."
            ws['D8'] = "中華全佑有限公司"
            ws['E8'] = "會計部 收"
    else:
        ws['C6'] = ''
        ws['D6'] = ''
        ws['D8'] = ''
        ws['E8'] = ''
        
        if not is_return:
            ws['K6'] = zip_code
            ws['L6'] = addr
            ws['L8'] = party_data.get('name') or ''
            ws['M8'] = "會計部 收"
        else:
            ws['I5'] = f"\n         {party_data.get('name') or ''}\n\n      {addr}"
            ws['K6'] = "350"
            ws['L6'] = "苗栗縣竹南鎮中美里10鄰保安林52-10號。"
            ws['L8'] = "中華全佑有限公司"
            ws['M8'] = "會計部 收"
            
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_address_label_excel(parties: list) -> bytes:
    wb = openpyxl.load_workbook("backend/templates/envelope_template.xlsx")
    ws = wb["地址標籤貼紙"]
    
    # Fill up to 16 labels in 2 columns
    for idx, p in enumerate(parties[:16]):
        row_idx = (idx // 2) * 5 + 2
        col_offset = 0 if (idx % 2 == 0) else 6
        
        name_cell = ws.cell(row=row_idx, column=col_offset + 2)
        addr_cell = ws.cell(row=row_idx + 2, column=col_offset + 2)
        
        name_cell.value = p.get('name') or ''
        addr_cell.value = p.get('address') or ''
        
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_purchase_order_excel(order_data: dict) -> bytes:
    # Simulates purchase printing, returning simple Excel bytes
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "進貨單"
    ws['A1'] = "中華全佑有限公司 - 進貨單"
    ws['A3'] = f"進貨單號: {order_data.get('purchase_id')}"
    ws['A4'] = f"日期: {order_data.get('date')}"
    ws['A5'] = f"廠商: {order_data.get('vendor_name')}"
    
    ws['A7'] = "品名規格花紋"
    ws['B7'] = "數量"
    ws['C7'] = "單價"
    ws['D7'] = "金額"
    
    items = order_data.get('items', [])
    for idx, item in enumerate(items):
        r = 8 + idx
        brand = item.get('brand') or ''
        spec = item.get('spec') or ''
        pattern = item.get('pattern') or ''
        ws[f'A{r}'] = f"{brand} {spec} {pattern}".strip()
        ws[f'B{r}'] = item.get('qty', 0)
        ws[f'C{r}'] = item.get('price', 0)
        ws[f'D{r}'] = item.get('qty', 0) * item.get('price', 0)
        
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_quote_excel(quote_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報價單"
    ws['A1'] = "中華全佑有限公司 - 報價單"
    ws['A3'] = f"報價單號: {quote_data.get('quote_id')}"
    ws['A4'] = f"類型: {quote_data.get('quote_type')}"
    ws['A5'] = f"對象: {quote_data.get('party_name')}"
    
    ws['A7'] = "品品項描述"
    ws['B7'] = "數量"
    ws['C7'] = "單價"
    ws['D7'] = "金額"
    
    items = quote_data.get('items', [])
    for idx, item in enumerate(items):
        r = 8 + idx
        brand = item.get('brand') or ''
        spec = item.get('spec') or ''
        pattern = item.get('pattern') or ''
        ws[f'A{r}'] = f"{brand} {spec} {pattern}".strip()
        ws[f'B{r}'] = item.get('qty', 0)
        ws[f'C{r}'] = item.get('price', 0)
        ws[f'D{r}'] = item.get('qty', 0) * item.get('price', 0)
        
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
