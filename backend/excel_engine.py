import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

def create_sales_order_excel(order_data: dict) -> bytes:
    wb = openpyxl.load_workbook("backend/templates/sales_order_template.xlsx")
    ws = wb["出貨單˙收據"]
    
    # Date parsing
    date_val = order_data.get('date', '')
    if isinstance(date_val, str) and date_val:
        try:
            date_str = datetime.fromisoformat(date_val.replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            date_str = date_val
    elif isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = ''
        
    cust_name = order_data.get('customer_name') or ''
    plate_num = order_data.get('plate_number') or ''
    created_by = order_data.get('created_by') or ''
    cust_addr = order_data.get('customer_address') or ''
    
    # Left side general info
    ws['D31'] = cust_name
    ws['D32'] = plate_num
    ws['J31'] = date_str
    ws['G32'] = created_by
    ws['D34'] = cust_addr
    
    # Right side general info
    ws['R31'] = cust_name
    ws['R32'] = plate_num
    ws['X31'] = date_str
    ws['U32'] = created_by
    ws['R34'] = cust_addr
    
    items = order_data.get('items', [])
    for idx, item in enumerate(items):
        if idx >= 8:  # Template supports up to 8 rows natively
            break
        r = 36 + idx
        brand = item.get('brand') or ''
        spec = item.get('spec') or item.get('specification') or ''
        pattern = item.get('pattern') or ''
        brand_spec_pattern = f"{brand} {spec} {pattern}".strip()
        service_type = item.get('service_type') or ''
        tire_pos = item.get('tire_position') or ''
        qty = item.get('qty') or 0
        price = item.get('price') or 0
        amount = qty * price
        note = item.get('note') or ''
        
        # Left side item row
        ws[f'B{r}'] = idx + 1
        ws[f'C{r}'] = order_data.get('order_id') or ''
        ws[f'D{r}'] = service_type
        ws[f'E{r}'] = brand_spec_pattern
        ws[f'G{r}'] = tire_pos
        ws[f'H{r}'] = note
        ws[f'I{r}'] = qty
        ws[f'J{r}'] = price
        ws[f'K{r}'] = amount
        
        # Right side item row
        ws[f'P{r}'] = idx + 1
        ws[f'Q{r}'] = order_data.get('order_id') or ''
        ws[f'R{r}'] = service_type
        ws[f'S{r}'] = brand_spec_pattern
        ws[f'U{r}'] = tire_pos
        ws[f'V{r}'] = ''
        ws[f'W{r}'] = qty
        ws[f'X{r}'] = price
        ws[f'Y{r}'] = amount

    # Totals
    subtotal = order_data.get('total_amount') or 0
    tax = order_data.get('tax_amount') or 0
    grand = order_data.get('grand_total') or 0
    
    # Left totals
    ws['K19'] = subtotal
    ws['H20'] = tax
    ws['K20'] = grand
    
    # Right totals
    ws['Y19'] = subtotal
    ws['V20'] = tax
    ws['Y20'] = grand
    
    # Save to memory
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_envelope_excel(party_data: dict, is_return: bool = False) -> bytes:
    """Generate Envelope layout (信封袋列印) matching the original format by loading template."""
    wb = openpyxl.load_workbook("backend/templates/envelope_template.xlsx")
    ws = wb["信封袋列印"]
    
    is_vendor = "vendor_id" in party_data
    
    # Extract zip code
    addr = party_data.get('address') or ''
    zip_code = ''
    if addr:
        import re
        m = re.search(r'\d{3}', addr)
        if m:
            zip_code = m.group(0)
        else:
            zip_code = addr[:3] if addr[:3].isdigit() else ''

    if not is_vendor:
        # Customer envelope (Columns B-E)
        ws['K6'] = ''
        ws['L6'] = ''
        ws['L8'] = ''
        ws['M8'] = ''
        
        if not is_return:
            ws['C6'] = zip_code
            ws['D6'] = addr
            ws['D8'] = party_data.get('name') or ''
            ws['E8'] = "會計部 收"
        else:
            ws['B5'] = f"\n         {party_data.get('name') or ''}\n\n      {addr}"
            ws['C6'] = "350"
            ws['D6'] = "苗栗縣竹南鎮中美里10鄰保安林52之10號."
            ws['D8'] = "中華全佑有限公司"
            ws['E8'] = "會計部 收"
    else:
        # Vendor envelope (Columns I-M)
        ws['C6'] = ''
        ws['D6'] = ''
        ws['D8'] = ''
        ws['E8'] = ''
        
        if not is_return:
            ws['K6'] = zip_code
            ws['L6'] = addr
            ws['L8'] = party_data.get('name') or ''
            ws['M8'] = "會計部 收"
        else:
            ws['I5'] = f"\n         {party_data.get('name') or ''}\n\n      {addr}"
            ws['K6'] = "350"
            ws['L6'] = "苗栗縣竹南鎮中美里10鄰保安林52-10號。"
            ws['L8'] = "中華全佑有限公司"
            ws['M8'] = "會計部 收"
            
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_address_label_excel(parties: list) -> bytes:
    """Generate Address Labels layout (地址貼列印)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "地址貼列印"

    font_large = Font(size=12, bold=True)
    
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20

    row_idx = 2
    for party in parties:
        ws.cell(row=row_idx, column=1, value="收件人 :").font = font_large
        ws.cell(row=row_idx, column=2, value=party.get('address', '')[:3] if party.get('address') else '')
        ws.cell(row=row_idx, column=3, value=party.get('address', '')).font = font_large
        
        ws.cell(row=row_idx+1, column=3, value=party.get('name', '')).font = font_large
        ws.cell(row=row_idx+1, column=4, value="會計部 收").font = font_large
        
        row_idx += 4 # Skip some rows for next label

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_purchase_order_excel(order_data: dict) -> bytes:
    """Generate a Purchase Order (進貨單) mimicking the original Excel layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "進貨單"

    title_font = Font(size=20, bold=True)
    header_font = Font(size=12, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    ws.merge_cells('A1:F1')
    ws['A1'] = "中華全佑 ERP - 進貨單"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = f"單號: {order_data.get('purchase_id', '')}"
    
    date_val = order_data.get('date', '')
    if isinstance(date_val, str) and date_val:
        try:
            date_str = datetime.fromisoformat(date_val.replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            date_str = date_val
    elif isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = ''
    ws['D3'] = f"日期: {date_str}"

    ws['A4'] = f"廠商: {order_data.get('vendor_name', '')}"
    ws['D4'] = f"統編: {order_data.get('vendor_uniform_number', '')}"
    ws['A5'] = f"電話: {order_data.get('vendor_phone', '')}"
    ws['A6'] = f"地址: {order_data.get('vendor_address', '')}"

    headers = ["商品編碼", "品名/規格/花紋", "部位/維修", "數量", "單價", "金額"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row_idx = 9
    for item in order_data.get('items', []):
        ws.cell(row=row_idx, column=1, value=item.get('product_code', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=f"{item.get('brand','')} {item.get('spec','')} {item.get('pattern','')}").border = thin_border
        ws.cell(row=row_idx, column=3, value=f"{item.get('tire_position','')} {item.get('service_type','')}").border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get('qty', 0)).border = thin_border
        ws.cell(row=row_idx, column=5, value=item.get('price', 0)).border = thin_border
        ws.cell(row=row_idx, column=6, value=item.get('amount', 0)).border = thin_border
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=5, value="未稅金額:").font = header_font
    ws.cell(row=row_idx, column=6, value=order_data.get('total_amount', 0))
    row_idx += 1
    ws.cell(row=row_idx, column=5, value="稅額:").font = header_font
    ws.cell(row=row_idx, column=6, value=order_data.get('tax_amount', 0))
    row_idx += 1
    ws.cell(row=row_idx, column=5, value="總計:").font = title_font
    ws.cell(row=row_idx, column=6, value=order_data.get('grand_total', 0)).font = title_font

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def create_quote_excel(order_data: dict) -> bytes:
    """Generate a Quote (報價單) mimicking the original Excel layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報價單"

    title_font = Font(size=20, bold=True)
    header_font = Font(size=12, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    ws.merge_cells('A1:F1')
    ws['A1'] = "中華全佑 ERP - 報價單"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = f"單號: {order_data.get('quote_id', '')}"
    
    date_val = order_data.get('date', '')
    if isinstance(date_val, str) and date_val:
        try:
            date_str = datetime.fromisoformat(date_val.replace('Z', '+00:00')).strftime('%Y-%m-%d')
        except:
            date_str = date_val
    elif isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = ''
    ws['D3'] = f"日期: {date_str}"

    ws['A4'] = f"客戶/廠商: {order_data.get('party_name', '')}"
    ws['D4'] = f"統編: {order_data.get('uniform_number', '')}"
    ws['A5'] = f"電話: {order_data.get('phone', '')}"
    ws['D5'] = f"有效期限: {order_data.get('valid_until', '')}"
    ws['A6'] = f"地址: {order_data.get('address', '')}"

    headers = ["商品編碼", "品名/規格/花紋", "部位/維修", "數量", "單價", "金額"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row_idx = 9
    for item in order_data.get('items', []):
        ws.cell(row=row_idx, column=1, value=item.get('product_code', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=f"{item.get('brand','')} {item.get('spec','')} {item.get('pattern','')}").border = thin_border
        ws.cell(row=row_idx, column=3, value=f"{item.get('tire_position','')} {item.get('service_type','')}").border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get('qty', 0)).border = thin_border
        ws.cell(row=row_idx, column=5, value=item.get('price', 0)).border = thin_border
        ws.cell(row=row_idx, column=6, value=item.get('amount', 0)).border = thin_border
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=5, value="未稅金額:").font = header_font
    ws.cell(row=row_idx, column=6, value=order_data.get('total_amount', 0))
    row_idx += 1
    ws.cell(row=row_idx, column=5, value="稅額:").font = header_font
    ws.cell(row=row_idx, column=6, value=order_data.get('tax_amount', 0))
    row_idx += 1
    ws.cell(row=row_idx, column=5, value="總計:").font = title_font
    ws.cell(row=row_idx, column=6, value=order_data.get('grand_total', 0)).font = title_font

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
